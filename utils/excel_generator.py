"""
Excel generation module with formatting and styling
Exports ONLY 3 columns as per Final Plan: key, value, comments
"""

import pandas as pd
from io import BytesIO
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def generate_excel(
    df: pd.DataFrame, 
    include_header: bool = True, 
    auto_width: bool = True
) -> BytesIO:
    """
    Generate Excel file from DataFrame with formatting
    ONLY exports 3 columns: key, value, comments (as per Final Plan)
    
    Args:
        df: DataFrame to export (must have key, value, comments columns)
        include_header: Whether to include column headers
        auto_width: Whether to auto-adjust column widths
        
    Returns:
        BytesIO: Excel file buffer
    """
    # Ensure only 3 required columns are exported
    required_columns = ['key', 'value', 'comments']
    export_df = df[[col for col in required_columns if col in df.columns]].copy()
    
    output = BytesIO()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Failed to get active worksheet"
    ws.title = "Extracted Data"
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_font = Font(name='Arial', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Write data
    if include_header:
        # Write headers
        for col_idx, column in enumerate(export_df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Write data rows
        for row_idx, row_data in enumerate(export_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
    else:
        # Write data without headers
        for row_idx, row_data in enumerate(export_df.values, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
    
    # Auto-adjust column widths
    if auto_width:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter  # type: ignore
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    if include_header:
        ws.freeze_panes = 'A2'
    
    # Save to buffer
    wb.save(output)
    output.seek(0)
    
    return output


def generate_excel_with_xlsxwriter(
    df: pd.DataFrame,
    include_header: bool = True,
    auto_width: bool = True
) -> BytesIO:
    """
    Alternative Excel generation using xlsxwriter for more features
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Extracted Data', index=False, header=include_header)
        
        workbook = writer.book
        worksheet = writer.sheets['Extracted Data']
        
        # Define formats
        header_format = workbook.add_format({  # type: ignore
            'bold': True,
            'font_color': 'white',
            'bg_color': '#366092',
            'border': 1,
            'align': 'center',
            'valign': 'vcenter',
            'text_wrap': True
        })
        
        cell_format = workbook.add_format({  # type: ignore
            'border': 1,
            'align': 'left',
            'valign': 'top',
            'text_wrap': True
        })
        
        # Apply header formatting
        if include_header:
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
        
        # Auto-adjust column widths
        if auto_width:
            for idx, col in enumerate(df.columns):
                max_length = max(
                    df[col].astype(str).map(len).max(),
                    len(col)
                ) + 2
                max_length = min(max_length, 50)
                worksheet.set_column(idx, idx, max_length, cell_format)
        
        # Freeze header row
        if include_header:
            worksheet.freeze_panes(1, 0)
    
    output.seek(0)
    return output


def generate_multi_sheet_excel(
    dataframes: dict,
    include_header: bool = True,
    auto_width: bool = True
) -> BytesIO:
    """
    Generate Excel file with multiple sheets
    
    Args:
        dataframes: Dictionary where keys are sheet names and values are DataFrames
        include_header: Whether to include headers
        auto_width: Whether to auto-adjust widths
        
    Returns:
        BytesIO: Excel file buffer
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        for sheet_name, df in dataframes.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=include_header)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Define formats
            header_format = workbook.add_format({  # type: ignore
                'bold': True,
                'font_color': 'white',
                'bg_color': '#366092',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            # Apply formatting
            if include_header:
                for col_num, value in enumerate(df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
            
            # Auto-adjust columns
            if auto_width:
                for idx, col in enumerate(df.columns):
                    max_length = max(
                        df[col].astype(str).map(len).max(),
                        len(col)
                    ) + 2
                    worksheet.set_column(idx, idx, min(max_length, 50))
    
    output.seek(0)
    return output
